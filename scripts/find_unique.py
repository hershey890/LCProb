"""Not for user use. For my personal development of this repo only.

Parses problems_raw.xlsx into problems_unique.csv which is used by lcprob.py
"""
from pathlib import Path
from collections import OrderedDict
import re
import openpyxl
import pandas as pd


def _get_problem_lists(file: Path | str) -> dict:
    """Open each workbook sheet into a dataframe and store in a dictionary"""
    prob_lists = OrderedDict()

    wb = openpyxl.load_workbook(file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        nums = []
        names = []
        diffs = []
        links = []
        lst = []

        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            cell = row[0]
            if row[0].value is None:
                break

            matches = re.search(
                r"\s*(\d+)\. (.+)(Easy|Med\.|Hard)", row[0].value
            )
            if matches is None:
                raise RuntimeError(
                    f"Problem {row[0].value} not parsed correctly"
                )

            nums.append(matches.group(1))
            names.append(matches.group(2))
            if matches.group(3) == "Med.":
                diffs.append("Medium")
            else:
                diffs.append(matches.group(3))
            link = re.match(r"(https://leetcode.com/problems/[\w-]+)", cell.hyperlink.target).group(1)
            links.append(link + "/")
            lst.append(sheet)

        data = pd.DataFrame(
            {"Number": nums, "Name": names, "Difficulty": diffs, "Link": links, "List": lst}
        )
        prob_lists[sheet] = data

    return prob_lists


def _get_unique_problems(prob_lists: dict) -> pd.DataFrame:
    """Get the unique problems from each sheet not not in previous sheets"""
    problems = []
    probs_per_sheet = OrderedDict()

    for key, df in prob_lists.items():
        probs = set(df.Number)
        unique_probs = probs - set(problems)
        problems += list(unique_probs)
        probs_per_sheet[key] = unique_probs
        prob_lists[key] = df[df.Number.isin(unique_probs)]
        n_col = len(prob_lists[key].columns)
        prob_lists[key].insert(n_col, "Date Completed", "")
        prob_lists[key].insert(n_col+1, "Completed", 0)

    return prob_lists


def _create_unique_csv(out_file: Path | str , prob_lists: dict) -> None:
    """Saves a .csv file with all unique problems from each sheet
    """
    out_df_list = []
    for key, df in prob_lists.items():
        # df["List"] = key
        out_df_list.append(df)

    out_df = pd.concat(out_df_list, ignore_index=True)
    out_df.to_csv(out_file, index=False)


def main() -> None:
    in_file = Path(__file__).parent.parent / Path("data/problems_raw.xlsx")
    out_file = Path(__file__).parent.parent / Path("data/problems_unique.csv")

    prob_lists = _get_problem_lists(in_file)
    assert len(prob_lists) == 3
    assert len(prob_lists["Grind75"]) == 75
    assert len(prob_lists["Grind169"]) == 169
    assert len(prob_lists["Neetcode150"]) == 150

    prob_lists = _get_unique_problems(prob_lists)
    assert len(prob_lists["Grind75"]) == 75
    assert len(prob_lists["Grind169"]) == 94
    assert len(prob_lists["Neetcode150"]) == 44

    _create_unique_csv(out_file, prob_lists)


if __name__ == "__main__":
    main()